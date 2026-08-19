# -*- coding: utf-8 -*-
"""KTRT 词库「同根词」修复/丰富管线（可复用于任意英语单词书）。

背景：早期 AI 填充把「同义词的派生词」误当「同根词」塞进同根词列（GRE 曾出错得离谱）。
本脚本把整套方法固化：
  1) 规则初筛（免费）：真同根词保留；同义词的派生词回迁到同义词列并标注为
     `原词（变形）`（如 abridge（abridgement, abridger））；重复/无关词删除。
  2) AI 补全（DeepSeek，可换厂商）：规则后仍不足 2 个同根词的词，按新定义补齐。
  3) 过滤：词典存在性（ECDICT）、同/反义词混入、简单屈折重复、生造词复核。
  4) 写回 Excel（可选同步 ktrt.db）。

适用性：任意英语单词书 Excel，只要含【单词】列即可；【音标】【词性释义】【同义词】
【同根词】【List】缺列时会自动补列（空列 + 全量 AI 补全），其余列原样保留。
内置备份：每次运行自动生成 _bak_时间戳.xlsx，可随时回滚。

用法（在 C:\\KTRT 项目内，用项目 venv 运行）：
  venv\\Scripts\\python.exe scripts\\fix_wordbook_roots.py ^
      --xlsx 任意词书.xlsx --tag 标识 [--book-id DB书id] [--rules-only]
  --book-id  : 同步 ktrt.db 中对应书（按 list_no+seq 匹配）；缺省只改 Excel。
  --rules-only : 只跑规则初筛（不调用 AI，用于预算/预览）。
中途可断点续跑：AI 批结果按批落盘，重跑自动跳过已完成的批次。
"""

import argparse
import io
import json
import os
import re
import shutil
import sqlite3
import sys
import time
import unicodedata
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)
from backend import ai  # noqa: E402

BASE = os.path.dirname(os.path.abspath(__file__))
DICT_DB = os.path.join(ROOT, "data", "dictionary.db")
KTRT_DB = os.path.join(ROOT, "data", "ktrt.db")

STD_HEADERS = ["【单词】", "【音标】", "【词性释义】", "【搭配】", "【短语】", "【同义词】",
               "【反义词】", "【同根词】", "【List】", "【语言】", "【单词书】"]

SUFFIXES = ["ationally", "isation", "ization", "iveness", "ability", "ibility", "ification",
            "ation", "ition", "sion", "tion", "ment", "ness", "er", "or", "ist", "ism", "ity",
            "ous", "ive", "able", "ible", "al", "ally", "ize", "ise", "ify", "ant", "ance",
            "ancy", "ence", "ency", "ure", "ary", "ory", "ing", "ed", "ies", "es", "ly", "s", "y", "e"]
PREFIXES = ["counter", "inter", "super", "trans", "under", "after", "over", "out", "pre", "pro",
            "post", "anti", "non", "mis", "dis", "un", "in", "im", "re", "de", "en", "be", "up",
            "ab", "ad", "com", "con", "ex", "per", "sub", "a"]


def stems(w):
    w = w.lower()
    out = {w}
    if len(w) <= 3:
        return out

    def add_variants(s):
        if len(s) < 3:
            return
        out.add(s)
        if s.endswith("c"):
            out.add(s + "t")
        if s.endswith("s"):
            out.add(s[:-1] + "d")
        if s.endswith("d"):
            out.add(s[:-1] + "s")
        if s.endswith("i"):
            out.add(s[:-1] + "y")
        if s.endswith("e"):
            out.add(s[:-1])
        if len(s) >= 4 and s[-1] == s[-2] and s[-1] in "bcdfglmnprstz":
            out.add(s[:-1])

    if w.endswith("ation") or w.endswith("ition"):
        add_variants(w[:-3])
    elif w.endswith("sion"):
        add_variants(w[:-3])
    elif w.endswith("tion"):
        add_variants(w[:-4])
    else:
        for suf in SUFFIXES:
            if w.endswith(suf) and len(w) - len(suf) >= 3:
                add_variants(w[:-len(suf)])
                break
    if w.endswith("e") and len(w) > 3:
        out.add(w[:-1])
    return out


def related(a, b):
    a, b = a.lower(), b.lower()
    if a == b:
        return True
    if stems(a) & stems(b):
        return True
    lcp = 0
    for x, y in zip(a, b):
        if x != y:
            break
        lcp += 1
    mn = min(len(a), len(b))
    return lcp >= 5 and lcp >= 0.6 * mn


def related_with_prefix(a, b):
    if related(a, b):
        return True
    for p in PREFIXES:
        if a.startswith(p) and len(a) - len(p) >= 3 and related(a[len(p):], b):
            return True
        if b.startswith(p) and len(b) - len(p) >= 3 and related(a, b[len(p):]):
            return True
    return False


def norm(w):
    return "".join(c for c in unicodedata.normalize("NFKD", w.lower()) if not unicodedata.combining(c))


def split(v):
    return [x.strip() for x in (v or "").split("；") if x.strip()]


def en_head(t):
    m = re.match(r"^[A-Za-z][A-Za-z\-\']*", t)
    return m.group(0).lower() if m else None


def clean_syn(syn_text):
    entries = split(syn_text)
    heads = [en_head(e) for e in entries]
    out = []
    for i, e in enumerate(entries):
        m = re.match(r"^(.*?)（(.*)）$", e)
        if m:
            base = m.group(1).strip()
            vars_list = [v.strip() for v in m.group(2).split(",") if v.strip()]
            keep = []
            for v in vars_list:
                vh = en_head(v)
                if vh and any(h == vh and bi != i for bi, h in enumerate(heads)):
                    continue
                keep.append(v)
            out.append("%s（%s）" % (base, ", ".join(keep)) if keep else base)
        else:
            out.append(e.strip())
    s = "；".join(out)
    s = re.sub(r"[（(]\s+", "（", s)
    s = re.sub(r",\s+", ", ", s)
    return s


def extract_json(text):
    text = re.sub(r"```(?:json)?", "", text).strip()
    a, b = text.find("{"), text.rfind("}")
    if a == -1 or b == -1:
        raise ValueError("no JSON in response")
    return json.loads(text[a:b + 1])


def load_dict_words():
    d = sqlite3.connect(DICT_DB)
    words = {r[0].strip().lower() for r in d.execute("SELECT word FROM dict")}
    d.close()
    return words


def triage(word, syns, roots):
    syn_heads = [(s, en_head(s)) for s in syns]
    keep, annot, drop = [], {}, []
    for t in roots:
        h = en_head(t)
        if not h:
            drop.append(t)
            continue
        if h == word.lower():
            continue
        if related(h, word):
            keep.append(t)
            continue
        base = None
        for s, sh in syn_heads:
            if sh and related(h, sh):
                base = s
                break
        if base:
            if h == en_head(base):
                drop.append(t)
            else:
                annot.setdefault(base, []).append(t)
        else:
            drop.append(t)
    seen = set()
    keep2 = []
    for t in keep:
        h = en_head(t)
        if h and h not in seen:
            seen.add(h)
            keep2.append(t)
    new_syns = []
    for s, sh in syn_heads:
        new_syns.append("%s（%s）" % (s, ", ".join(annot[s])) if s in annot else s)
    return keep2, "；".join(new_syns), drop


def build_refill_prompt(part):
    lines = []
    for w in part:
        meaning = (w["meaning"] or "").strip().replace("\n", " ")
        lines.append("%s | %s | known=%s" % (w["w"], meaning, w["root"] or "无"))
    return (
        "你是英语词汇编辑，专精词源学与构词法。\n\n"
        "【同根词定义】同根词 = 与目标词共享同一词根、由该词根加前后缀派生的词族成员。"
        "例：act → action, active, activity, actor；abbreviate → abbreviation, abbreviator。\n\n"
        "【严格禁止】不得填入：1) 近义词或反义词（例：abbreviate 的同根词不得写 abridge / contract / shorten）；"
        "2) 词根不同的词（例：aberrant 的同根词不得写 deviance / deviant / deviate）；"
        "3) 目标词本身；4) 纯屈折变化 -s / -ed / -ing（除非是常用重要派生词）。\n\n"
        "【要求】每个词输出 2-4 个真实存在的标准英语同根词；确实不足则写 1 个。"
        "给出的 known= 为已验证正确的同根词，必须包含在输出中。\n\n"
        "只输出一个 JSON 对象：键为单词本身，值为同根词字符串数组。不要输出任何其他文字，不要 markdown 代码块。\n\n"
        "词表（格式：单词 | 词性释义 | known）：\n" + "\n".join(lines)
    )


def run_ai_batches(ai_words, out_dir, prompt_builder, batch=75, max_tokens=8192, temperature=0.2):
    os.makedirs(out_dir, exist_ok=True)
    existing = {fn for fn in os.listdir(out_dir) if fn.startswith("batch_") and fn.endswith(".json")}
    result = {}
    t0 = time.time()
    for i in range(0, len(ai_words), batch):
        part = ai_words[i:i + batch]
        out_fn = os.path.join(out_dir, "batch_%03d.json" % (i // batch))
        if os.path.basename(out_fn) in existing:
            result.update(json.load(open(out_fn, encoding="utf-8")))
            print("[skip] %s" % os.path.basename(out_fn))
            continue
        merged = {}
        for attempt in range(3):
            try:
                content = ai.chat([{"role": "user", "content": prompt_builder(part)}],
                                  max_tokens=max_tokens, temperature=temperature)
                merged = extract_json(content)
                break
            except Exception as e:
                print("  retry %s (%d): %s" % (os.path.basename(out_fn), attempt + 1, str(e)[:140]))
                time.sleep(4)
        with open(out_fn, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False)
        got = sum(1 for w in part if w["w"] in merged)
        result.update(merged)
        print("batch %d: got %d/%d, elapsed %.0fs" % (i // batch, got, len(part), time.time() - t0))
    return result


def filter_ai_items(word, items, syn_set, ant_set, base_families, dict_words):
    added = []
    for item in items:
        x = (item or "").strip().lower().strip(".,; ")
        nx = norm(x)
        if not re.match(r"^[a-z][a-z\-']*$", x):
            continue
        if nx == norm(word):
            continue
        if related_with_prefix(x, word):
            pass
        elif nx in syn_set or nx in ant_set:
            continue
        elif x not in dict_words:
            continue
        stem = None
        if x.endswith("ing") and len(x) > 4:
            stem = x[:-3]
        elif x.endswith("ed") and len(x) > 3:
            stem = x[:-2]
        elif x.endswith("es") and len(x) > 3:
            stem = x[:-2]
        elif x.endswith("s") and len(x) > 3 and not x.endswith(("ss", "us", "is", "ous", "ness", "less")):
            stem = x[:-1]
        if stem and len(stem) >= 3 and any(related(stem, b) for b in base_families):
            continue
        if nx not in {norm(b) for b in base_families}:
            base_families.add(x)
            added.append(x)
    return added


def rare_verify(kept_not_dict, out_dir):
    items = sorted({en_head(x) for x in kept_not_dict if en_head(x) and "-" not in en_head(x)})
    if not items:
        return {}
    verdict = {}
    os.makedirs(out_dir, exist_ok=True)
    existing = {fn for fn in os.listdir(out_dir) if fn.startswith("batch_") and fn.endswith(".json")}
    for i in range(0, len(items), 100):
        part = items[i:i + 100]
        out_fn = os.path.join(out_dir, "batch_%03d.json" % (i // 100))
        if os.path.basename(out_fn) in existing:
            verdict.update(json.load(open(out_fn, encoding="utf-8")))
            continue
        prompt = (
            "判断下面每个英文单词是否为真实存在的英语单词（包括罕见词、古词、专业术语、合法派生词）。"
            "符合以下任一情况判 false：拼写错误、AI 生造词、纯外来语原词（法语/意大利语等未英语化词）、网络俚语生造词。"
            "只输出一个 JSON 对象：{\"单词\": true 或 false}，不要任何其他文字。\n词表：\n" + "、".join(part)
        )
        merged = {}
        for attempt in range(3):
            try:
                content = ai.chat([{"role": "user", "content": prompt}], max_tokens=4096, temperature=0.1)
                merged = extract_json(content)
                break
            except Exception as e:
                print("retry %s: %s" % (os.path.basename(out_fn), str(e)[:120]))
                time.sleep(4)
        with open(out_fn, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False)
        verdict.update(merged)
    return verdict


def ensure_columns(ws, header):
    """补齐标准列（缺列追加到表头，值为空），返回列名->索引。"""
    for name in STD_HEADERS:
        if name not in header:
            ws.cell(row=1, column=len(header) + 1, value=name)
            header.append(name)
    return {h: i for i, h in enumerate(header)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="任意英语单词书 Excel（至少含【单词】列）")
    ap.add_argument("--tag", required=True, help="标识（用于中间产物目录名）")
    ap.add_argument("--book-id", type=int, default=None, help="ktrt.db 中对应书 id（缺省只改 Excel）")
    ap.add_argument("--rules-only", action="store_true", help="只跑规则初筛，不调用 AI")
    args = ap.parse_args()

    xlsx = args.xlsx
    tag = args.tag
    plan_path = os.path.join(BASE, tag + "_plan.json")
    refill_dir = os.path.join(BASE, tag + "_refill")
    rare_dir = os.path.join(BASE, tag + "_rare")
    result_path = os.path.join(BASE, tag + "_result.json")
    dict_words = load_dict_words()

    bak = xlsx.replace(".xlsx", "_bak_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S"))
    shutil.copy2(xlsx, bak)
    print("备份:", bak)

    import openpyxl

    # 补列（写入模式）
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    ix = ensure_columns(ws, header)
    wb.save(xlsx)
    wb.close()

    # 规则初筛
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb.worksheets[0]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(header)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    plan = []
    stats = Counter()
    seq_by = Counter()
    for row in rows:
        w = (row[ix["【单词】"]] or "").strip()
        if not w:
            continue
        lst = row[ix["【List】"]] or 1
        seq_by[lst] += 1
        seq = seq_by[lst]
        meaning = row[ix["【词性释义】"]] or ""
        roots = split(row[ix["【同根词】"]])
        syns = split(row[ix["【同义词】"]])
        keep, syn_new, drop = triage(w, syns, roots)
        ai_flag = len(keep) < 2
        plan.append({
            "w": w, "list": lst, "seq": seq, "meaning": meaning,
            "syn": syn_new, "root": "；".join(keep),
            "ai": ai_flag, "had_roots": bool(roots), "dropped": "；".join(drop),
        })
        if ai_flag:
            stats["ai"] += 1
        if not keep:
            stats["empty_after"] += 1

    json.dump(plan, open(plan_path, "w", encoding="utf-8"), ensure_ascii=False)
    print("规则初筛完成: 词条=%d | 需AI补全=%d | 规则后空=%d" % (len(plan), stats["ai"], stats["empty_after"]))

    if args.rules_only:
        print("仅规则模式，未写回（预览请查看 plan: %s）" % plan_path)
        return

    # AI 补全
    ai_words = [p for p in plan if p["ai"]]
    print("AI 补全词数:", len(ai_words))
    refill = run_ai_batches(ai_words, refill_dir, build_refill_prompt)

    # 合并 + 过滤 + 写回
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(header)}
    result = []
    kept_not_dict = []
    seq_by2 = Counter()
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        cells = {h: row[j] for j, h in enumerate(header)}
        w = (cells["【单词】"].value or "").strip()
        if not w:
            continue
        lst = cells["【List】"].value or 1
        seq_by2[lst] = seq_by2.get(lst, 0) + 1
        seq = seq_by2[lst]
        p = plan[i]
        syn_new = clean_syn(p["syn"])
        roots_new = split(p["root"])
        if p["ai"]:
            aiv = refill.get(w)
            if isinstance(aiv, list):
                syn_set = {norm(s) for s in split(syn_new)}
                ant_set = {norm(s) for s in split((cells["【反义词】"].value or ""))}
                base_families = {w}
                for r in roots_new:
                    eh = en_head(r)
                    if eh:
                        base_families.add(eh)
                roots_new += filter_ai_items(w, aiv, syn_set, ant_set, base_families, dict_words)
        seen = set()
        final = []
        for r in roots_new:
            key = norm(r)
            if key and key not in seen:
                seen.add(key)
                final.append(r)
        final = final[:4]
        cells["【同义词】"].value = syn_new
        cells["【同根词】"].value = "；".join(final)
        for r in final:
            eh = en_head(r)
            if eh and eh not in dict_words:
                kept_not_dict.append(eh)
        result.append({"w": w, "list": lst, "seq": seq, "syn": syn_new, "root": "；".join(final)})
    wb.save(xlsx)
    print("Excel 已写回:", xlsx)

    # 稀有词 AI 复核（省钱：只复核不在词典里的）
    verdict = rare_verify(kept_not_dict, rare_dir)
    removed = {w for w, v in verdict.items() if v is not True}
    for r in result:
        keep = []
        for x in r["root"].split("；"):
            h = en_head(x)
            if h and (h in removed or "-" in h):
                continue
            keep.append(x)
        r["root"] = "；".join(keep)

    json.dump(result, open(result_path, "w", encoding="utf-8"), ensure_ascii=False)

    wb = openpyxl.load_workbook(xlsx)
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(header)}
    res_by = {(r["list"], r["seq"]): r["root"] for r in result}
    seq_by3 = Counter()
    n = 0
    for row in ws.iter_rows(min_row=2):
        w = (row[ix["【单词】"]].value or "").strip()
        if not w:
            continue
        lst = row[ix["【List】"]].value or 1
        seq_by3[lst] = seq_by3.get(lst, 0) + 1
        seq = seq_by3[lst]
        newv = res_by.get((lst, seq), "")
        if (row[ix["【同根词】"]].value or "") != newv:
            row[ix["【同根词】"]].value = newv
            n += 1
    wb.save(xlsx)
    print("复核后写回行数:", n)

    if args.book_id:
        c = sqlite3.connect(KTRT_DB)
        upd = 0
        for r in result:
            cur = c.execute(
                "UPDATE words SET synonyms=?, root_words=? WHERE book_id=? AND list_no=? AND seq=?",
                (r["syn"], r["root"], args.book_id, r["list"], r["seq"]),
            )
            upd += cur.rowcount
        c.commit()
        c.close()
        print("DB(book_id=%d) 更新行数: %d" % (args.book_id, upd))

    dist = Counter(len(r["root"].split("；")) if r["root"] else 0 for r in result)
    print("最终根数分布:", sorted(dist.items()))
    print("空根词数:", sum(1 for r in result if not r["root"]))
    print("完成。原始备份: %s" % bak)


if __name__ == "__main__":
    main()
