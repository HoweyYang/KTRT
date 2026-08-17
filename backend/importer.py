import csv
import os
import re

from . import db

WORD_KEYS = {'【单词】', 'word', '单词', 'word_en', 'entry'}
PHON_KEYS = {'【音标】', 'phonetic', '音标', 'pronunciation', 'ipa'}
MEAN_KEYS = {'【词性释义】', 'meaning', '释义', '词性释义', 'translation', '中文', 'definition_cn'}
COLL_KEYS = {'【搭配】', 'collocations', '搭配'}
PHR_KEYS = {'【短语】', 'phrases', '短语'}
SYN_KEYS = {'【同义词】', 'synonyms', '同义词'}
ANT_KEYS = {'【反义词】', 'antonyms', '反义词'}
ROOT_KEYS = {'【同根词】', 'root_words', '同根词', 'word_family'}
LIST_KEYS = {'【List】', 'list', 'list_no', '单元', 'unit'}
LANG_KEYS = {'【语言】', 'language', '语言'}
BOOK_KEYS = {'【单词书】', 'book', '单词书', 'bookname'}


def _clean(v):
    return '' if v is None else str(v).strip()


def _colmap(header):
    m = {}
    for i, h in enumerate(header):
        h = _clean(h).lower()
        for key, keys in (
            ('word', WORD_KEYS), ('phonetic', PHON_KEYS), ('meaning', MEAN_KEYS),
            ('collocations', COLL_KEYS), ('phrases', PHR_KEYS), ('synonyms', SYN_KEYS),
            ('antonyms', ANT_KEYS), ('root_words', ROOT_KEYS), ('list_no', LIST_KEYS),
            ('language', LANG_KEYS), ('book_name', BOOK_KEYS),
        ):
            if h in {k.lower() for k in keys} and key not in m:
                m[key] = i
    return m


def _num(v):
    digits = re.sub(r'\D', '', _clean(v))
    return int(digits) if digits else 1


def _build_rows(records, cmap):
    rows = []
    for rec in records:
        word = _clean(rec[cmap['word']]) if 'word' in cmap and cmap['word'] < len(rec) else ''
        if not word:
            continue
        rows.append({
            'word': word,
            'phonetic': _clean(rec[cmap['phonetic']]) if 'phonetic' in cmap and cmap['phonetic'] < len(rec) else '',
            'meaning': _clean(rec[cmap['meaning']]) if 'meaning' in cmap and cmap['meaning'] < len(rec) else '',
            'collocations': _clean(rec[cmap['collocations']]) if 'collocations' in cmap and cmap['collocations'] < len(rec) else '',
            'phrases': _clean(rec[cmap['phrases']]) if 'phrases' in cmap and cmap['phrases'] < len(rec) else '',
            'synonyms': _clean(rec[cmap['synonyms']]) if 'synonyms' in cmap and cmap['synonyms'] < len(rec) else '',
            'antonyms': _clean(rec[cmap['antonyms']]) if 'antonyms' in cmap and cmap['antonyms'] < len(rec) else '',
            'root_words': _clean(rec[cmap['root_words']]) if 'root_words' in cmap and cmap['root_words'] < len(rec) else '',
            'list_no': _num(rec[cmap['list_no']]) if 'list_no' in cmap and cmap['list_no'] < len(rec) else 1,
        })
    return rows


def parse_xlsx(path):
    from openpyxl import load_workbook  # 延迟导入，降低启动内存
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [_clean(c) for c in next(it)]
    cmap = _colmap(header)
    if 'word' not in cmap:
        raise ValueError('Excel 缺少单词列；请包含【单词】或 word 表头')
    records = list(it)
    book_name = ''
    language = ''
    if 'book_name' in cmap:
        for rec in records:
            if cmap['book_name'] < len(rec):
                v = _clean(rec[cmap['book_name']])
                if v:
                    book_name = v
                    break
    if 'language' in cmap:
        for rec in records:
            if cmap['language'] < len(rec):
                v = _clean(rec[cmap['language']])
                if v:
                    language = v
                    break
    return book_name, language, _build_rows(records, cmap)


def parse_csv(path):
    rows_all = []
    with open(path, encoding='utf-8-sig', newline='') as f:
        for row in csv.reader(f):
            rows_all.append(row)
    if not rows_all:
        return '', '', []
    header = [_clean(c) for c in rows_all[0]]
    cmap = _colmap(header)
    has_word_header = 'word' in cmap
    records = rows_all[1:] if has_word_header else rows_all
    book_name = ''
    language = ''
    if 'book_name' in cmap:
        for rec in records:
            if cmap['book_name'] < len(rec):
                v = _clean(rec[cmap['book_name']])
                if v:
                    book_name = v
                    break
    if 'language' in cmap:
        for rec in records:
            if cmap['language'] < len(rec):
                v = _clean(rec[cmap['language']])
                if v:
                    language = v
                    break
    if 'word' not in cmap:
        # 无表头：每行第一个非空单元格作为纯单词
        rows = []
        for rec in records:
            for cell in rec:
                w = _clean(cell)
                if w:
                    rows.append({'word': w, 'phonetic': '', 'meaning': '', 'collocations': '',
                                 'phrases': '', 'synonyms': '', 'antonyms': '', 'root_words': '',
                                 'list_no': 1})
                    break
        return book_name, language, rows
    return book_name, language, _build_rows(records, cmap)


def parse_txt(path):
    rows = []
    with open(path, encoding='utf-8-sig') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.lower().startswith(('word', '单词')):
                continue
            if '\t' in line:
                parts = [p.strip() for p in line.split('\t')]
            elif ',' in line:
                parts = [p.strip() for p in line.split(',')]
            else:
                parts = [line]
            while len(parts) < 3:
                parts.append('')
            rows.append({
                'word': parts[0],
                'phonetic': parts[1],
                'meaning': parts[2],
                'collocations': '', 'phrases': '', 'synonyms': '', 'antonyms': '',
                'root_words': '', 'list_no': 1,
            })
    return '', '', rows


def parse_file(path, forced_book='', forced_language=''):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        book_name, language, rows = parse_xlsx(path)
    elif ext == '.csv':
        book_name, language, rows = parse_csv(path)
    else:
        book_name, language, rows = parse_txt(path)
    book_name = (forced_book or book_name or os.path.splitext(os.path.basename(path))[0]).strip() or '新单词书'
    language = (forced_language or language or '英语').strip()
    if not rows:
        raise ValueError('文件中没有有效词条')
    return book_name, language, rows


def import_book(path, forced_book='', forced_language=''):
    book_name, language, rows = parse_file(path, forced_book, forced_language)
    with db._lock:
        with db.get_conn() as conn:
            cur = conn.execute(
                'INSERT OR IGNORE INTO word_books(name, language, source) VALUES(?,?,?)',
                (book_name, language, path),
            )
            book_id = cur.lastrowid
            if cur.rowcount == 0:
                book_id = conn.execute(
                    'SELECT id FROM word_books WHERE name=?', (book_name,)
                ).fetchone()['id']
            seq_counter = {}
            insert_rows = []
            for r in rows:
                seq_counter[r['list_no']] = seq_counter.get(r['list_no'], 0) + 1
                insert_rows.append((
                    book_id, r['list_no'], seq_counter[r['list_no']],
                    r['word'], r['phonetic'], r['meaning'], r['collocations'],
                    r['phrases'], r['synonyms'], r['antonyms'], r['root_words'],
                ))
            conn.executemany(
                'INSERT OR REPLACE INTO words(book_id, list_no, seq, word, phonetic, meaning, '
                'collocations, phrases, synonyms, antonyms, root_words) VALUES(?,?,?,?,?,?,?,?,?,?,?)',
                insert_rows,
            )
    return {'book_name': book_name, 'language': language, 'rows': len(rows)}
